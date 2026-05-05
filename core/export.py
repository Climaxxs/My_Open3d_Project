import csv
import os
from datetime import datetime
import numpy as np


class MeasurementExporter:

    @staticmethod
    def export_to_csv(file_path, table_data, measurement_results=None):
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)

                # 写入标题
                writer.writerow(['家畜点云尺寸测量系统 - 测量报告'])
                writer.writerow([f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
                writer.writerow([])

                # 写入测量结果摘要（如果有）
                if measurement_results:
                    writer.writerow(['测量结果摘要'])
                    writer.writerow(['项目', '数值 (m)'])
                    for label, value in measurement_results.items():
                        writer.writerow([label, f'{value:.3f}'])
                    writer.writerow([])

                # 写入表格数据
                writer.writerow(['测量记录明细'])
                writer.writerow(['编号', '类型', '数值 (m)', '单位', '时间'])
                for row in table_data:
                    writer.writerow(row)

                writer.writerow([])
                writer.writerow([f'共 {len(table_data)} 条记录'])

            return True, "导出成功"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def export_to_xlsx(file_path, table_data, measurement_results=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "测量报告"

            # 标题
            ws.merge_cells('A1:E1')
            ws['A1'] = '家畜点云尺寸测量系统 - 测量报告'
            ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:E2')
            ws['A2'] = f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws['A2'].alignment = Alignment(horizontal='center')

            row = 4

            # 测量结果摘要
            if measurement_results:
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = '测量结果摘要'
                ws[f'A{row}'].font = Font(name='微软雅黑', size=13, bold=True)
                row += 1

                ws[f'A{row}'] = '项目'
                ws[f'B{row}'] = '数值 (m)'
                for col in ['A', 'B']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                row += 1

                for label, value in measurement_results.items():
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = f'{value:.3f}'
                    row += 1

                row += 1

            # 测量记录明细
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = '测量记录明细'
            ws[f'A{row}'].font = Font(name='微软雅黑', size=13, bold=True)
            row += 1

            # 表头
            headers = ['编号', '类型', '数值 (m)', '单位', '时间']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center')
            row += 1

            # 数据
            for data_row in table_data:
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = f'共 {len(table_data)} 条记录'
            ws[f'A{row}'].font = Font(italic=True)

            # 调整列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 15

            wb.save(file_path)
            return True, "导出成功"

        except ImportError:
            return False, "需要安装openpyxl库: pip install openpyxl"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    @staticmethod
    def export_to_txt(file_path, table_data, measurement_results=None):
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write('=' * 50 + '\n')
                f.write('家畜点云尺寸测量系统 - 测量报告\n')
                f.write(f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                f.write('=' * 50 + '\n\n')

                if measurement_results:
                    f.write('【测量结果摘要】\n')
                    f.write('-' * 30 + '\n')
                    max_label_len = max(len(label) for label in measurement_results.keys())
                    for label, value in measurement_results.items():
                        f.write(f'{label:<{max_label_len + 2}} {value:.3f} m\n')
                    f.write('\n')

                f.write('【测量记录明细】\n')
                f.write('-' * 30 + '\n')
                f.write(f'{"编号":<6} {"类型":<12} {"数值":<10} {"单位":<6} {"时间":<10}\n')
                f.write('-' * 30 + '\n')

                for row in table_data:
                    f.write(f'{row[0]:<6} {row[1]:<12} {row[2]:<10} {row[3]:<6} {row[4]:<10}\n')

                f.write('\n' + '-' * 30 + '\n')
                f.write(f'共 {len(table_data)} 条记录\n')

            return True, "导出成功"
        except Exception as e:
            return False, f"导出失败: {str(e)}"